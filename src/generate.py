#!/usr/bin/env python3
"""
Cover Whale PDF Generator  v3
─────────────────────────────
Reads company data from Excel → fills Cover Whale insurance policy template
→ outputs one ready PDF per company.

Run:  py generate.py
"""

import fitz
import openpyxl
import os, sys, urllib.request, zipfile, io, logging, random, calendar, inspect, re
from datetime import datetime, date, timedelta
from pathlib import Path
from PIL import Image, ImageFilter, ImageEnhance
import numpy as np

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# ─── LOGGING ─────────────────────────────────────────────────────────────────

LOG_DIR = Path(__file__).parent.parent / "log"
LOG_DIR.mkdir(exist_ok=True)

logger = logging.getLogger("coverwhale")
if not logger.handlers:
    logger.setLevel(logging.INFO)
    fh = logging.FileHandler(LOG_DIR / "coverwhale.log", encoding="utf-8")
    fh.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", datefmt="%Y-%m-%d %H:%M:%S"))
    logger.addHandler(fh)

# ─── CONFIG ──────────────────────────────────────────────────────────────────

SCRIPT_DIR      = Path(__file__).parent
PROJECT_DIR     = SCRIPT_DIR.parent
ASSETS_DIR      = PROJECT_DIR / "assets"
TEMPLATE_PDF    = ASSETS_DIR / "template" / "Cover Whale -   VIATIC LLC.pdf"
EXCEL_FILE      = ASSETS_DIR / "companies.xlsx"
OUTPUT_DIR      = PROJECT_DIR / "output"
STARTING_POLICY = "CUS09116674"

FONT_REG  = ASSETS_DIR / "DejaVuSansCondensed.ttf"
FONT_BOLD = ASSETS_DIR / "DejaVuSansCondensed-Bold.ttf"
FONT_ZIP  = "https://sourceforge.net/projects/dejavu/files/dejavu/2.37/dejavu-fonts-ttf-2.37.zip/download"

# Arial fonts — macOS system path, falling back to the Windows system path.
_MAC_ARIAL   = Path("/System/Library/Fonts/Supplemental/Arial.ttf")
_MAC_ARIAL_B = Path("/System/Library/Fonts/Supplemental/Arial Bold.ttf")
ARIAL_REG  = _MAC_ARIAL   if _MAC_ARIAL.exists()   else Path("C:/Windows/Fonts/arial.ttf")
ARIAL_BOLD = _MAC_ARIAL_B if _MAC_ARIAL_B.exists() else Path("C:/Windows/Fonts/arialbd.ttf")


def _sys_font(*candidates, fallback):
    """First existing path among `candidates`, else `fallback`.

    Lets a document name the system font it wants on both macOS and Windows
    without the module failing to import on whichever machine lacks it.
    """
    for c in candidates:
        if Path(c).exists():
            return Path(c)
    return fallback

# Confirmation-of-Coverage policy term. The start day is randomised within this
# month and the term runs to the same day one year later. Shifting the term means
# changing these two values and nothing else — tests/test_coc_dates.py reads them
# rather than repeating the month and year.
COC_TERM_YEAR  = 2026
COC_TERM_MONTH = 4                      # April


def _coc_dates(rng=None):
    """Dates for the Confirmation-of-Coverage doc. Start = random day in
    COC_TERM_MONTH of COC_TERM_YEAR; term end = same day one year later;
    due = start + 21 days.
    Returns the exact format strings used at each spot in the template."""
    rng = rng or random
    day = rng.randint(1, calendar.monthrange(COC_TERM_YEAR, COC_TERM_MONTH)[1])
    # A Feb 29 start has no counterpart in a non-leap year; fall back to the
    # last day of the month, as policy terms conventionally do.
    end_day = min(day, calendar.monthrange(COC_TERM_YEAR + 1, COC_TERM_MONTH)[1])
    start = date(COC_TERM_YEAR,     COC_TERM_MONTH, day)
    end   = date(COC_TERM_YEAR + 1, COC_TERM_MONTH, end_day)
    due = start + timedelta(days=21)
    return {
        "start_slash": start.strftime("%m/%d/%Y"),
        "end_slash": end.strftime("%m/%d/%Y"),
        "start_dash": start.strftime("%m-%d-%Y"),
        "end_dash": end.strftime("%m-%d-%Y"),
        "due": f"{due.month}/{due.day}/{due.year}",
    }


# ─── UTILITY BILL CONFIG ─────────────────────────────────────────────────────
UTILITY_TEMPLATE = ASSETS_DIR / "template" / "Utility_IVORY JULIUS CHRISTOPHER.pdf"
UT_NAME    = "IVORY JULIUS CHRISTOPHER"
UT_ADDR1   = "10318 CHEEVES,"
UT_ADDR2   = "HOUSTON, TX 77016"

# ─── NGANGA CONFIG (Motor Carrier Liability Declaration Page) ────────────────
# Professional Transportation Risk Retention Group — 2-page declaration.
# Every replaced field on this template is Calibri / Calibri-Bold at 11.04pt.
NGANGA_TEMPLATE = ASSETS_DIR / "template" / "Pricely Fane LLC_Policy_040226_unlocked-1-2.pdf"
NGANGA_POLICY   = "PT-26042618-01"

# Calibri ships with Windows and with Microsoft Office on macOS; fall back to
# Arial (metrically closer than DejaVu) so a machine without it still renders.
CALIBRI_REG  = _sys_font("/Library/Fonts/Microsoft/Calibri.ttf",
                         "/Library/Fonts/Calibri.ttf",
                         str(Path.home() / "Library/Fonts/Calibri.ttf"),
                         "C:/Windows/Fonts/calibri.ttf", fallback=ARIAL_REG)
CALIBRI_BOLD = _sys_font("/Library/Fonts/Microsoft/Calibri Bold.ttf",
                         "/Library/Fonts/Calibrib.ttf",
                         str(Path.home() / "Library/Fonts/Calibri Bold.ttf"),
                         "C:/Windows/Fonts/calibrib.ttf", fallback=ARIAL_BOLD)

NG_COMPANY = "Pricely Fane LLC"
NG_ADDR1   = "5270 Millenia Blvd"
NG_ADDR2   = "Orlando, FL 32839-5636"
NG_VIN     = "3AKJGLD50GSGY1387"
NG_YEAR    = "2016"
NG_DRIVER  = "Pricely Gracius"
NG_DOB     = "10/3/1998"
NG_STATE   = "FL"
NG_LICENSE = "G611610943000"
NG_PREPARED = "5/7/2026"      # footer date — its own span, label is separate
NG_PERIOD   = "5/7/2026 12:01 AM to 5/7/2027 12:00 AM"

NG_FONTSIZE     = 11.04   # every replaced field on this template
NG_RIGHT_EDGE   = 540.0   # visible right edge the two header lines align to
NG_LOGO_RIGHT_X = 230.0   # letterhead logo bbox ends x222.7 — header must clear it
NG_BOX_RIGHT_X  = 536.0   # inner right edge of the page-1 value boxes (border 539.5)

# ─── CONFIRMATION OF COVERAGE CONFIG ─────────────────────────────────────────
COC_TEMPLATE = ASSETS_DIR / "template" / "1 GUY 1 GIRL 1 TRUCK LLC new insurance.pdf"
COC_NAME_P1  = "SAFE ROAD FREIGHT INC"        # page 1 cover-sheet sample company
COC_NAME     = "1 GUY 1 GIRL 1 TRUCK LLC"     # pages 2 & 5 sample company
COC_ADDR1    = "1234 N KENILWORTH AVE"        # page 2 mailing address line 1
COC_ADDR2    = "OAK PARK, IL 60302"           # page 2 mailing address line 2

# Exact fonts/sizes/colors read from the template's own text spans:
#   page 1  company = Arial-BoldMT 22pt #171717 ; date = Arial-Black 16pt #171717 (centered)
#   pages 2-4 values = DejaVuSans 10pt black ; page 5 values = DejaVuSans 9pt black
COC_FONT_BODY  = ASSETS_DIR / "DejaVuSans.ttf"  # pages 2-5 values
COC_FONT_TITLE = _sys_font("/System/Library/Fonts/Supplemental/Arial Bold.ttf",
                           "C:/Windows/Fonts/arialbd.ttf", fallback=COC_FONT_BODY)
COC_FONT_DATE1 = _sys_font("/System/Library/Fonts/Supplemental/Arial Black.ttf",
                           "C:/Windows/Fonts/ariblk.ttf", fallback=COC_FONT_BODY)
COC_CLR_DARK  = (0x17 / 255.0, 0x17 / 255.0, 0x17 / 255.0)  # #171717
COC_CLR_BLACK = (0.0, 0.0, 0.0)


def _coc_set(page, old, new, *, size, color, font, center=False):
    """Cover the template's sample text with white and write the new value at
    the same baseline using an exact font/size/color — no heuristics, no
    background sampling (these fields all sit on white)."""
    fp = str(font)
    fobj = fitz.Font(fontfile=fp)
    fname = "F" + str(abs(hash(fp)) % 100000)
    rects = page.search_for(old)
    if not rects:
        return

    # Strip the sample text before painting; writing first would redact the
    # new value along with it. See _strip_text.
    _strip_text(page, rects)

    for rect in rects:
        page.draw_rect(fitz.Rect(rect.x0 - 1, rect.y0 - 1, rect.x1 + 1, rect.y1 + 1),
                       color=(1, 1, 1), fill=(1, 1, 1), width=0)
        tw = fobj.text_length(new, fontsize=size)
        x = (page.rect.width - tw) / 2 if center else rect.x0
        y = rect.y1 - 1.0
        page.insert_text((x, y), new, fontfile=fp, fontname=fname, fontsize=size, color=color)

# ─── TEMPLATE VALUES (VIATIC LLC base PDF) ───────────────────────────────────
T_COMPANY = "VIATIC LLC"
T_USDOT   = "USDOT # 3846659"
T_ADDR1   = "3975 NW 176TH ST"
T_ADDR2   = "MIAMI GARDENS, FL 33055"
T_POLICY  = "CUS09114581"

# ─── STATIC TEMPLATE VALUES (Tr=3 invisible in template → must be re-rendered) ─
T_FROM      = "From:"
T_TO        = "To:"
T_TERM_FROM = "October 14, 2025"
T_TERM_TO   = "October 14, 2026"
T_BROKER1   = "Empire State Brokerage Services LLC - DAVID SCHEPSMAN"
T_BROKER2   = "(DAVID@ESBSLLC.COM )"
T_ISSUED_TM = "11:31:12 EST (Eastern Standard Time)"

# ─── PAGE 2 COLUMN CENTERS (Confirmation of Coverage table) ──────────────────
# Center of Mailing Address column (~x400 to x560) → used for centering the value
P2_MAILING_ADDR_CX  = 480.0

# ─── TOP-RIGHT FIXED RIGHT EDGE ──────────────────────────────────────────────
# Both policy # and company name right-align to the same edge across all pages
TOP_RIGHT_X = 552.0

# ─── TEXT COLORS ─────────────────────────────────────────────────────────────
CLR_HEADER  = (54/255, 54/255, 54/255)    # #363636 — top-right header text
CLR_TITLE   = (51/255, 51/255, 51/255)    # #333333 — center bold title / USDOT

# ─── FONTS ───────────────────────────────────────────────────────────────────

def ensure_fonts():
    if FONT_REG.exists() and FONT_BOLD.exists():
        return
    print("  Downloading DejaVu fonts (one-time) ...")
    with urllib.request.urlopen(FONT_ZIP) as r:
        data = r.read()
    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        for m in zf.namelist():
            if m.endswith("DejaVuSansCondensed.ttf") and not FONT_REG.exists():
                FONT_REG.write_bytes(zf.read(m));  print(f"  + {FONT_REG.name}")
            elif m.endswith("DejaVuSansCondensed-Bold.ttf") and not FONT_BOLD.exists():
                FONT_BOLD.write_bytes(zf.read(m)); print(f"  + {FONT_BOLD.name}")
    if not FONT_REG.exists() or not FONT_BOLD.exists():
        print("  Font extraction failed — place TTF files next to generate.py")
        sys.exit(1)

# ─── HELPERS ─────────────────────────────────────────────────────────────────

def increment_policy(p: str) -> str:
    digits = "".join(c for c in p if c.isdigit())
    prefix = "".join(c for c in p if not c.isdigit())
    return f"{prefix}{int(digits)+1:0{len(digits)}d}"


def increment_nganga_policy(p: str) -> str:
    """
    Increment only the middle block of a PT-style policy number.
    "PT-26042618-01" -> "PT-26042619-01"

    increment_policy() cannot be used here: it strips every digit in the
    string, so the trailing "-01" would be folded into the counter and
    "PT-26042618-01" would become "PT--2604261802".
    """
    parts = p.split("-")
    if len(parts) != 3 or not parts[1].isdigit():
        raise ValueError(f"Unexpected policy format: {p!r} (want PT-NNNNNNNN-NN)")
    parts[1] = f"{int(parts[1])+1:0{len(parts[1])}d}"
    return "-".join(parts)


# ─── RANDOMISED FIELD GENERATORS (Nganga declaration page) ───────────────────

VIN_LETTERS = "ABCDEFGHJKLMNPRSTUVWXYZ"   # I, O and Q never appear in a VIN

VIN_YEAR_CODE = {2014: "E", 2015: "F", 2016: "G",
                 2017: "H", 2018: "J", 2019: "K"}   # VIN position 10

_VIN_TRANS = {**{str(d): d for d in range(10)},
              "A": 1, "B": 2, "C": 3, "D": 4, "E": 5, "F": 6, "G": 7, "H": 8,
              "J": 1, "K": 2, "L": 3, "M": 4, "N": 5, "P": 7, "R": 9,
              "S": 2, "T": 3, "U": 4, "V": 5, "W": 6, "X": 7, "Y": 8, "Z": 9}
_VIN_WEIGHTS = [8, 7, 6, 5, 4, 3, 2, 10, 0, 9, 8, 7, 6, 5, 4, 3, 2]


def vin_check_digit(vin17: str) -> str:
    """Standard NHTSA check digit for position 9. Returns '0'-'9' or 'X'."""
    total = sum(_VIN_TRANS[c] * w for c, w in zip(vin17.upper(), _VIN_WEIGHTS))
    r = total % 11
    return "X" if r == 10 else str(r)


def random_year(lo: int = 2014, hi: int = 2019, rng=None) -> int:
    return (rng or random).randint(lo, hi)


def random_vin(year: int, wmi: str = "3AK", rng=None) -> str:
    """
    Build a valid-looking VIN for the given model year.

    Mirrors the character-class pattern of the template's VIN
    (3AKJGLD50GSGY1387):  3AK  LLLL D  [check]  [year]  L  LL DDDD

    The WMI stays fixed so the VIN keeps agreeing with the Make column
    ("Freightliner"), position 10 encodes the year so it cannot contradict the
    Year column, and position 9 is a real check digit.
    """
    rng = rng or random
    if year not in VIN_YEAR_CODE:
        raise ValueError(f"No VIN year code for {year} (supported: {sorted(VIN_YEAR_CODE)})")

    vds   = "".join(rng.choice(VIN_LETTERS) for _ in range(4)) + str(rng.randint(0, 9))
    plant = rng.choice(VIN_LETTERS)
    seq   = "".join(rng.choice(VIN_LETTERS) for _ in range(2)) + \
            "".join(str(rng.randint(0, 9)) for _ in range(4))

    vin = f"{wmi}{vds}0{VIN_YEAR_CODE[year]}{plant}{seq}"   # '0' = check placeholder
    return vin[:8] + vin_check_digit(vin) + vin[9:]


def random_dob(min_age: int = 25, max_age: int = 49, today: date = None, rng=None) -> str:
    """
    Random date of birth for someone between min_age and max_age years old.
    Formatted M/D/YYYY to match the template ("10/3/1998").
    """
    rng = rng or random
    today = today or date.today()
    newest = today - timedelta(days=int(min_age * 365.25))
    oldest = today - timedelta(days=int((max_age + 1) * 365.25) - 1)
    dob = oldest + timedelta(days=rng.randint(0, (newest - oldest).days))
    return f"{dob.month}/{dob.day}/{dob.year}"


# Driver-licence formats. 'N' = digit, 'L' = random letter,
# '#' = first letter of the driver's surname (states that encode it).
LICENSE_FORMATS = {
    "AL": "NNNNNNN",       "AK": "NNNNNNN",       "AZ": "LNNNNNNNN",
    "AR": "NNNNNNNNN",     "CA": "LNNNNNNN",      "CO": "NNNNNNNNN",
    "CT": "NNNNNNNNN",     "DE": "NNNNNNN",       "DC": "NNNNNNN",
    "FL": "#NNNNNNNNNNNN", "GA": "NNNNNNNNN",     "HI": "LNNNNNNNN",
    "ID": "LLNNNNNNL",     "IL": "#NNNNNNNNNNN",  "IN": "NNNNNNNNNN",
    "IA": "NNNNNNNNN",     "KS": "LNNNNNNNN",     "KY": "#NNNNNNNN",
    "LA": "NNNNNNNNN",     "ME": "NNNNNNN",       "MD": "#NNNNNNNNNNNN",
    "MA": "LNNNNNNNN",     "MI": "#NNNNNNNNNNNN", "MN": "#NNNNNNNNNNNN",
    "MS": "NNNNNNNNN",     "MO": "NNNNNNNNN",     "MT": "NNNNNNNNN",
    "NE": "LNNNNNNNN",     "NV": "NNNNNNNNNN",    "NH": "NNNNNNNNNN",
    "NJ": "#NNNNNNNNNNNNNN", "NM": "NNNNNNNNN",   "NY": "NNNNNNNNN",
    "NC": "NNNNNNNNN",     "ND": "NNNNNNNNN",     "OH": "LLNNNNNN",
    "OK": "LNNNNNNNNN",    "OR": "NNNNNNNNN",     "PA": "NNNNNNNN",
    "RI": "NNNNNNN",       "SC": "NNNNNNNNN",     "SD": "NNNNNNNN",
    "TN": "NNNNNNNNN",     "TX": "NNNNNNNN",      "UT": "NNNNNNNNN",
    "VT": "NNNNNNNN",      "VA": "#NNNNNNNN",     "WA": "#NNNNNNNNNNN",
    "WV": "NNNNNNN",       "WI": "#NNNNNNNNNNNNN", "WY": "NNNNNNNNN",
}


def random_state(rng=None) -> str:
    return (rng or random).choice(sorted(LICENSE_FORMATS))


def random_license(state: str, last_name: str = "", rng=None) -> str:
    """
    Build a licence number in the given state's format. Where the state encodes
    the holder's surname initial (FL, MI, MD, NJ, ...), that letter comes from
    last_name so the number agrees with the driver's name.
    """
    rng = rng or random
    fmt = LICENSE_FORMATS.get(state.upper())
    if not fmt:
        raise ValueError(f"No licence format for state {state!r}")

    initial = (last_name.strip()[:1] or rng.choice(VIN_LETTERS)).upper()
    if not initial.isalpha():
        initial = rng.choice(VIN_LETTERS)

    out = []
    for ch in fmt:
        if ch == "N":
            out.append(str(rng.randint(0, 9)))
        elif ch == "L":
            out.append(rng.choice("ABCDEFGHIJKLMNOPQRSTUVWXYZ"))
        elif ch == "#":
            out.append(initial)
    return "".join(out)


def split_address(addr: str):
    """
    Split address into (street, city_state_zip).
    Handles Excel cells where lines are separated by \n or \n\n.
    Falls back to comma split.
    """
    addr = addr.strip()
    if "\n" in addr:
        parts = [p.strip() for p in addr.split("\n") if p.strip()]
        return (parts[0], parts[-1]) if len(parts) >= 2 else (parts[0], "")
    parts = addr.split(",", 1)
    return (parts[0].strip(), parts[1].strip()) if len(parts) == 2 else (addr, "")


def height_to_params(h: float):
    """
    (fontsize, use_bold, center_on_page) from text bbox height.
    Calibrated from the Cover Whale template PDF.
    """
    if h > 14:    return 16.50, True,  True    # Page 1 large bold company title
    elif h > 11:  return 10.62, False, False   # Page 1 box name / address
    elif h > 7.5: return  8.56, False, False   # Page 2 Named Insured value
    else:         return  7.36, False, False   # Top-right tiny header labels


def sample_bg(pix, rect, pw, ph):
    """
    Sample the PDF background color just LEFT of the text bbox.
    Returns (r, g, b) floats 0-1.  Falls back to white if anything fails.
    """
    if pix is None:
        return (1.0, 1.0, 1.0)
    sx = pix.width  / pw
    sy = pix.height / ph
    px = int((rect.x0 - 4) * sx)
    py = int(((rect.y0 + rect.y1) / 2) * sy)
    if px < 0:
        px = int((rect.x1 + 4) * sx)
    px = max(0, min(pix.width  - 1, px))
    py = max(0, min(pix.height - 1, py))
    try:
        return tuple(c / 255.0 for c in pix.pixel(px, py)[:3])
    except Exception:
        return (1.0, 1.0, 1.0)


def merge_hits(hits, x_gap=4.0):
    """
    Merge search_for() rects that are fragments of a single run of text.

    search_for splits a match wherever glyph heights change — in a date like
    "5/7/2026" the slashes are taller than the digits, so one match comes back
    as 8 rects. Left unmerged the replacement is drawn once per fragment.
    """
    if not hits:
        return []

    def joins(a, b):
        return (min(a.y1, b.y1) - max(a.y0, b.y0) > 0 and
                b.x0 <= a.x1 + x_gap and a.x0 <= b.x1 + x_gap)

    groups = []
    for r in sorted(hits, key=lambda r: r.x0):
        for i, g in enumerate(groups):
            if joins(g, r):
                groups[i] = g | r
                break
        else:
            groups.append(fitz.Rect(r))

    # Unioning can bring two groups into contact, so consolidate until stable.
    merged = True
    while merged:
        merged = False
        for i in range(len(groups)):
            for j in range(i + 1, len(groups)):
                if joins(groups[i], groups[j]):
                    groups[i] |= groups.pop(j)
                    merged = True
                    break
            if merged:
                break
    return groups


def _strip_text(page, rects):
    """Delete the text inside `rects` from the page's content stream.

    Painting a rectangle over text only hides it: the characters stay in the
    text layer and come straight back out of copy/paste or any text extractor,
    so a generated document still carried the template's sample company, address
    and dates. Redaction actually removes them.

    Images and line art are explicitly preserved. The defaults
    (images=PDF_REDACT_IMAGE_REMOVE, graphics=PDF_REDACT_LINE_ART_REMOVE_IF_TOUCHED)
    delete anything a redaction rect touches, which wipes table rules, shaded
    rows and logos — that is what made an earlier attempt at this corrupt the
    page and get reverted to draw_rect.
    """
    if not rects:
        return
    for r in rects:
        # Shrink to the middle band of the line box. Redaction drops every
        # character whose box the rect touches, and glyph boxes span the full
        # line height including ascender and descender — so a full-height rect
        # also catches the lines directly above and below. On the CoC mailing
        # address that deleted the street line that had just been written.
        inset = min(r.height * 0.3, 2.5)
        page.add_redact_annot(fitz.Rect(r.x0, r.y0 + inset, r.x1, r.y1 - inset))
    page.apply_redactions(images=fitz.PDF_REDACT_IMAGE_NONE,
                          graphics=fitz.PDF_REDACT_LINE_ART_NONE)


def _plan_replacement(page, old_text, new_text,
                      fontsize=None, bold=False, center=False,
                      cell_center_x=None, cell_right_x=None, cell_left_x=None,
                      cell_bounds=None,
                      top_right_x=None,
                      x_min=None, x_max=None,
                      y_min=None, y_max=None,
                      font_reg=None, font_bold=None,
                      merge=False, fit_left_x=None, fit_right_x=None,
                      min_fontsize=6.5, **_painting_opts):
    """
    Work out where and how each occurrence of old_text will be rewritten.

    Returns a list of (rect, x, y, size, fontfile, use_bold, new_text) and
    changes nothing. Planning is split from painting so several fields can be
    stripped in one redaction pass before any of them is drawn.
    """
    hits = page.search_for(old_text)
    if not hits:
        return []
    if merge:
        hits = merge_hits(hits)

    pw = page.rect.width

    plans = []
    for rect in hits:
        # ── positional guards ────────────────────────────────────────────────
        if x_min is not None and rect.x0 < x_min: continue
        if x_max is not None and rect.x1 > x_max: continue
        if y_min is not None and rect.y0 < y_min: continue
        if y_max is not None and rect.y1 > y_max: continue

        # ── font / size ──────────────────────────────────────────────────────
        h = rect.y1 - rect.y0
        if fontsize is None:
            sz, use_bold, use_center = height_to_params(h)
        else:
            sz, use_bold, use_center = fontsize, bold, center

        # ── x alignment ──────────────────────────────────────────────────────
        f_reg  = font_reg  or FONT_REG
        f_bold = font_bold or FONT_BOLD
        fp       = str(f_bold if use_bold else f_reg)
        font_obj = fitz.Font(fontfile=fp)
        tw       = font_obj.text_length(new_text, fontsize=sz)

        # ── shrink to fit ────────────────────────────────────────────────────
        avail = None
        if fit_left_x is not None:
            right = (cell_right_x if cell_right_x is not None else
                     top_right_x  if top_right_x  is not None else rect.x1)
            avail = right - fit_left_x
        elif fit_right_x is not None:
            avail = fit_right_x - rect.x0
        if avail is not None and avail > 0 and tw > avail:
            sz = max(min_fontsize, sz * avail / tw)
            tw = font_obj.text_length(new_text, fontsize=sz)

        if use_center:
            # Large bold title — center across full page width
            x = (pw - tw) / 2
        elif cell_bounds is not None:
            # Smart center: center within cell, left-align with padding if too wide
            cb_left, cb_right = cell_bounds
            cell_w = cb_right - cb_left
            if tw < cell_w:
                x = cb_left + (cell_w - tw) / 2
            else:
                x = cb_left + 2.0
        elif cell_right_x is not None:
            # Right-align within a table column
            x = cell_right_x - tw
        elif cell_center_x is not None:
            # Center within a table column
            x = cell_center_x - tw / 2
        elif cell_left_x is not None:
            # Fixed left edge within a table column
            x = cell_left_x
        elif rect.x0 > 400 and rect.y0 < 80:
            # Top-right corner header → right-align to fixed common edge
            trx = top_right_x if top_right_x is not None else TOP_RIGHT_X
            x = min(trx, pw - 4) - tw
        else:
            # Body text — left-align at original position
            x = rect.x0

        y = rect.y1 - 1.0          # baseline just inside the bottom of the bbox
        plans.append((rect, x, y, sz, fp, use_bold, new_text))

    return plans


def _paint_plans(page, plans, pix, color=None, font_reg=None, fontname_tag=None,
                 cover=True, **_planning_opts):
    """Draw planned replacements. Must run after _strip_text, never before."""
    pw, ph = page.rect.width, page.rect.height

    for rect, x, y, sz, fp, use_bold, new_text in plans:
        # ── cover old text with a rectangle matching the actual background ──
        if cover:
            bg = sample_bg(pix, rect, pw, ph)
            cover_rect = fitz.Rect(rect.x0 - 1.0, rect.y0 - 1.0,
                                   rect.x1 + 1.0, rect.y1 + 1.0)
            page.draw_rect(cover_rect, color=bg, fill=bg, width=0)

        # ── write new text on top ─────────────────────────────────────────────
        if fontname_tag is not None:
            fnm = f"{fontname_tag}Bd" if use_bold else fontname_tag
        elif font_reg is not None:
            fnm = "ArialBd" if use_bold else "Arial"
        else:
            fnm = "DejaVuSCBd" if use_bold else "DejaVuSC"
        text_color = color if color is not None else (0, 0, 0)
        page.insert_text((x, y), new_text, fontfile=fp, fontname=fnm, fontsize=sz, color=text_color)


# Destination codepoints MuPDF picks that we never actually write, and what they
# should have been. Keys and values are the 4-hex-digit forms used in a CMap.
_TOUNICODE_FIXES = {b"00ad": b"002d",     # SOFT HYPHEN    -> HYPHEN-MINUS
                    b"00a0": b"0020"}     # NO-BREAK SPACE -> SPACE

# One `<src> <dst>` pair on its own line, i.e. a single-character bfchar entry.
# Deliberately not bfrange: rewriting the start of a range would shift every
# codepoint in it. Anchored so only the destination is ever touched — <00a0>
# also occurs on the left as a glyph code (`<00a0> <00e6>`), and a blind byte
# replacement would corrupt that entry.
_BFCHAR_ENTRY = re.compile(
    rb"^(\s*<[0-9A-Fa-f]{4}>\s*<)(" + b"|".join(_TOUNICODE_FIXES) + rb")(>\s*)$",
    re.MULTILINE | re.IGNORECASE)


def _fix_text_extraction(doc):
    """Correct the ToUnicode CMaps MuPDF generates for the fonts we embed.

    macOS Arial maps U+002D and U+00AD onto one hyphen glyph, and U+0020 and
    U+00A0 onto one space glyph. MuPDF's reverse lookup takes the higher
    codepoint, so the CMap it writes says

        <0010> <00ad>      hyphen glyph -> SOFT HYPHEN
        <0003> <00a0>      space  glyph -> NO-BREAK SPACE

    The page renders correctly either way, but extraction and copy/paste hand
    back "PT\\xad26042619\\xad01" — an invisible character inside a policy
    number — and every word separated by NBSP instead of a space.

    We never write a real soft hyphen or NBSP, so correcting those two
    destinations is unambiguous. Worst case a template's own font legitimately
    mapped one, and its extraction gains an ordinary hyphen or space instead.
    """
    for xref in range(1, doc.xref_length()):
        tu = doc.xref_get_key(xref, "ToUnicode")
        if tu[0] != "xref":
            continue
        stream_xref = int(tu[1].split()[0])
        try:
            cmap = doc.xref_stream(stream_xref)
        except Exception:
            continue
        if not cmap:
            continue
        fixed = _BFCHAR_ENTRY.sub(
            lambda m: m.group(1) + _TOUNICODE_FIXES[m.group(2).lower()] + m.group(3),
            cmap)
        if fixed != cmap:
            doc.update_stream(stream_xref, fixed)


def save_pdf(doc, out):
    """Save a generated document. Repairs text extraction first — every save
    goes through here so no output can skip that."""
    _fix_text_extraction(doc)
    doc.save(str(out), garbage=4, deflate=True)


def _accepted_opts(fn, skip):
    """Keyword names `fn` names explicitly, ignoring its catch-all **kwargs."""
    return {name for name, p in inspect.signature(fn).parameters.items()
            if name not in skip and p.kind is not p.VAR_KEYWORD}


# The union of what planning and painting accept. Derived from the signatures so
# it cannot drift as either side gains an option.
REPLACE_OPTS = (_accepted_opts(_plan_replacement, {"page", "old_text", "new_text"})
                | _accepted_opts(_paint_plans, {"page", "plans", "pix"}))


def _check_opts(kw, caller):
    """Reject unknown options.

    replace_on_page takes **kw and hands the same dict to both _plan_replacement
    and _paint_plans, each of which absorbs the other's keywords through a
    catch-all. That means a misspelling — colour= for color=, say — is accepted
    in silence and the option simply never applies. There are 30-odd call sites
    where colour and position are load-bearing, so a typo must fail loudly.
    """
    unknown = sorted(set(kw) - REPLACE_OPTS)
    if unknown:
        raise TypeError(
            f"{caller}() got unexpected keyword argument(s) {unknown}. "
            f"Accepted: {sorted(REPLACE_OPTS)}"
        )


def replace_on_page(page, old_text, new_text, pix=None, **kw):
    """
    Find every occurrence of old_text on page that passes the x/y filters,
    delete it, cover it with a filled rectangle matching the background, then
    write new_text at the correct position.

    Runs in two passes: every replacement is worked out first, the original
    text is stripped in a single redaction pass, and only then is anything
    painted. The new text has to be written after the redaction — writing it
    first would strip it along with the original.

    Keyword arguments beyond `pix` are split between _plan_replacement and
    _paint_plans; the notable additions are:

    fontname_tag  explicit CMap tag so each font family stays distinct; without
                  it any custom font is tagged "Arial", which collides once more
                  than one non-DejaVu family is embedded.
    merge         union hits that are fragments of one run. search_for splits a
                  match wherever glyph heights change, so a date like "5/7/2026"
                  comes back as 8 rects and would be replaced 8 times.
    fit_left_x    shrink until right-aligned text clears an obstacle on its left.
    fit_right_x   shrink until left-aligned text stays inside its table cell.
    cover         paint the background rectangle. Redaction already removes the
                  old text and preserves the shading under it, so a template
                  whose rows are closer together than their hit rects are tall
                  is better off with cover=False — the rectangle is what bleeds
                  into the neighbouring row, not the text.
    """
    _check_opts(kw, "replace_on_page")
    plans = _plan_replacement(page, old_text, new_text, **kw)
    if not plans:
        return
    _strip_text(page, [p[0] for p in plans])
    _paint_plans(page, plans, pix, **kw)


def replace_many(page, fields, pix=None, **common):
    """
    Replace several fields as a single unit.

    replace_on_page redacts and repaints within one call, so two calls that
    touch overlapping x-ranges on the same line will have the second call's
    redaction delete text the first call has already written. On the Nganga
    header that truncated the company name at exactly the x where the policy
    number's rect began. Planning every field first, stripping once, and only
    then painting removes the hazard.

    `fields` is a sequence of (old_text, new_text, overrides) triples; each
    overrides dict is merged over `common`.
    """
    _check_opts(common, "replace_many")
    batches = []
    for old, new, kw in fields:
        _check_opts(kw, f"replace_many field {old!r}")
        opts = {**common, **kw}
        plans = _plan_replacement(page, old, new, **opts)
        if plans:
            batches.append((plans, opts))

    if not batches:
        return

    _strip_text(page, [p[0] for plans, _ in batches for p in plans])
    for plans, opts in batches:
        _paint_plans(page, plans, pix, **opts)

# ─── PAGE FILL FUNCTIONS ─────────────────────────────────────────────────────

def fill_page1(page, company, usdot, addr1, addr2, policy, pix):
    """Page 1 — cover page."""

    # Top-right:  TGL Policy #:  CUS...  — right edge aligned to truck picture (x≈569)
    replace_on_page(page,
                    f"TGL Policy #:  {T_POLICY}",
                    f"TGL Policy #:  {policy}",
                    fontsize=7.36, top_right_x=569.0, pix=pix, color=CLR_HEADER)

    # Top-right company name
    replace_on_page(page, T_COMPANY, company,
                    fontsize=7.36, top_right_x=569.0, pix=pix, x_min=400, y_max=80, color=CLR_HEADER)

    # Centre bold company name (large title)
    replace_on_page(page, T_COMPANY, company,
                    fontsize=16.50, bold=True, center=True, pix=pix, x_max=400, y_min=80, y_max=400, color=CLR_TITLE)

    # Box name (below title)
    replace_on_page(page, T_COMPANY, company,
                    fontsize=10.62, pix=pix, x_max=400, y_min=400, y_max=600)

    # USDOT line (bold, centred)
    replace_on_page(page, T_USDOT, f"USDOT # {usdot}",
                    fontsize=11.0, bold=True, center=True, pix=pix, color=CLR_TITLE)

    # Address box — restrict to box area only (y 560–640)
    replace_on_page(page, T_ADDR1, addr1,
                    fontsize=10.62, pix=pix, y_min=560, y_max=610)
    replace_on_page(page, T_ADDR2, addr2,
                    fontsize=10.62, pix=pix, y_min=600, y_max=640)

    # Re-render static values (template stores them as Tr=3 invisible text)
    replace_on_page(page, T_FROM,      T_FROM,      pix=pix, fontsize=10.62, y_min=625, y_max=650)
    replace_on_page(page, T_TERM_FROM, T_TERM_FROM, pix=pix, fontsize=10.62, y_min=625, y_max=650)
    replace_on_page(page, T_TO,        T_TO,        pix=pix, fontsize=10.62, y_min=625, y_max=650)
    replace_on_page(page, T_TERM_TO,   T_TERM_TO,   pix=pix, fontsize=10.62, y_min=625, y_max=650)
    replace_on_page(page, T_BROKER1,   T_BROKER1,   pix=pix, fontsize=10.62, y_min=645, y_max=670)
    replace_on_page(page, T_BROKER2,   T_BROKER2,   pix=pix, fontsize=10.62, y_min=665, y_max=690)
    replace_on_page(page, T_TERM_FROM, T_TERM_FROM, pix=pix, fontsize=10.62, y_min=685, y_max=710)
    replace_on_page(page, T_ISSUED_TM, T_ISSUED_TM, pix=pix, fontsize=10.62, y_min=685, y_max=710)


def fill_page2(page, company, addr1, addr2, policy, pix):
    """Page 2 — Confirmation of Coverage."""

    # Top-right header
    replace_on_page(page,
                    f"TGL Policy #:  {T_POLICY}",
                    f"TGL Policy #:  {policy}",
                    fontsize=7.36, top_right_x=569.0, pix=pix, color=CLR_HEADER)

    # Top-right company name (right-aligned, top-right corner only)
    replace_on_page(page, T_COMPANY, company,
                    fontsize=7.36, top_right_x=569.0, pix=pix, x_min=400, y_max=80, color=CLR_HEADER)

    # Named Insured — smart centered within cell (42.8 – 217.9)
    replace_on_page(page, T_COMPANY, company,
                    fontsize=8.56,
                    cell_bounds=(42.8, 217.9),
                    pix=pix, x_max=300, y_min=200, y_max=240)

    # Policy Number — smart centered within cell (305.8 – 437.1)
    replace_on_page(page, T_POLICY, policy,
                    fontsize=9.63,
                    cell_bounds=(305.8, 437.1),
                    pix=pix, x_min=300, y_min=155, y_max=195)

    # Mailing Address — smart centered within cell (393.7 – 568.8)
    replace_on_page(page, T_ADDR1, addr1,
                    fontsize=8.56,
                    cell_bounds=(393.7, 568.8),
                    pix=pix, y_min=200, y_max=225)
    replace_on_page(page, T_ADDR2, addr2,
                    fontsize=8.56,
                    cell_bounds=(393.7, 568.8),
                    pix=pix, y_min=215, y_max=245)


def fill_page_header_only(page, company, policy, pix):
    """Pages 3+ — only top-right corner needs updating."""
    replace_on_page(page,
                    f"TGL Policy #:  {T_POLICY}",
                    f"TGL Policy #:  {policy}",
                    fontsize=7.36, top_right_x=569.0, pix=pix, color=CLR_HEADER)
    replace_on_page(page, T_COMPANY, company,
                    fontsize=7.36, top_right_x=569.0, pix=pix, x_min=400, y_max=80, color=CLR_HEADER)

# ─── UTILITY BILL FILL ────────────────────────────────────────────────────────

def fill_utility(page, company, addr1, addr2, pix):
    """
    Replace company name and address on page 1 of the Comcast utility template.
    Two locations: upper-left header area and bottom payment slip.
    """
    ar = dict(font_reg=ARIAL_REG, font_bold=ARIAL_BOLD)

    # Upper-left: bold company name (size 12)
    replace_on_page(page, UT_NAME, company,
                    fontsize=12.0, bold=True, pix=pix,
                    y_min=110, y_max=145, x_max=300, **ar)

    # Upper-left: address lines (size 9)
    replace_on_page(page, UT_ADDR1, addr1,
                    fontsize=9.0, pix=pix,
                    y_min=155, y_max=180, x_max=200, **ar)
    replace_on_page(page, UT_ADDR2, addr2,
                    fontsize=9.0, pix=pix,
                    y_min=170, y_max=195, x_max=200, **ar)

    # Bottom payment slip: company name (size 9, regular)
    replace_on_page(page, UT_NAME, company,
                    fontsize=9.0, pix=pix,
                    y_min=625, y_max=650, x_max=200, **ar)

    # Bottom payment slip: address lines (size 9)
    replace_on_page(page, UT_ADDR1, addr1,
                    fontsize=9.0, pix=pix,
                    y_min=640, y_max=660, x_max=200, **ar)
    replace_on_page(page, UT_ADDR2, addr2,
                    fontsize=9.0, pix=pix,
                    y_min=650, y_max=670, x_max=200, **ar)


def generate_utility(company: str, address: str, output_dir: Path = None) -> Path:
    """Generate a Comcast utility bill PDF with the given company name and address."""
    if output_dir is None:
        output_dir = OUTPUT_DIR
    output_dir.mkdir(exist_ok=True)

    addr1, addr2 = split_address(address.upper())
    company_up = company.strip().upper()

    doc = fitz.open(UTILITY_TEMPLATE)
    page = doc[0]
    pix = page.get_pixmap(dpi=72)
    fill_utility(page, company_up, addr1, addr2, pix)

    safe = (company_up
            .replace("/","-").replace("\\","-").replace(":","")
            .replace("*","").replace("?","").replace('"',"")
            .replace("<","").replace(">","").replace("|","")
            .replace("'",""))
    out = output_dir / f"Utility_{safe}.pdf"
    save_pdf(doc, out)
    doc.close()
    logger.info(f"Utility bill saved: {out.name}")
    return out


def generate_coc(company: str, address: str, output_dir: Path = None, dates: dict = None) -> Path:
    """Fill the 6-page Confirmation-of-Coverage template for the given company
    and address, with auto-generated policy dates. Output is trimmed to the
    first 6 pages (the trailing binder pages 7-12 are dropped)."""
    if output_dir is None:
        output_dir = OUTPUT_DIR
    output_dir.mkdir(exist_ok=True)

    addr1, addr2 = split_address(address.upper())
    company_up = company.strip().upper()
    d = dates or _coc_dates()

    doc = fitz.open(COC_TEMPLATE)

    # Page 1 — Arial cover, centered, #171717: company (Bold 22pt), term (Black 16pt).
    p = doc[0]
    _coc_set(p, COC_NAME_P1, company_up, size=22, color=COC_CLR_DARK,
             font=COC_FONT_TITLE, center=True)
    _coc_set(p, "10/16/2025 - 10/16/2026", f"{d['start_slash']} - {d['end_slash']}",
             size=16, color=COC_CLR_DARK, font=COC_FONT_DATE1, center=True)

    # Page 2 — Confirmation of Coverage: DejaVuSans 10pt black.
    p = doc[1]
    _coc_set(p, COC_NAME, company_up, size=10, color=COC_CLR_BLACK, font=COC_FONT_BODY)
    _coc_set(p, COC_ADDR1, addr1, size=10, color=COC_CLR_BLACK, font=COC_FONT_BODY)
    _coc_set(p, COC_ADDR2, addr2, size=10, color=COC_CLR_BLACK, font=COC_FONT_BODY)
    _coc_set(p, "10/16/2025", d["start_slash"], size=10, color=COC_CLR_BLACK, font=COC_FONT_BODY)
    _coc_set(p, "10/16/2026", d["end_slash"], size=10, color=COC_CLR_BLACK, font=COC_FONT_BODY)

    # Pages 3 & 4 — date only, DejaVuSans 10pt black.
    for i in (2, 3):
        _coc_set(doc[i], "10/16/2025", d["start_slash"], size=10, color=COC_CLR_BLACK, font=COC_FONT_BODY)

    # Page 5 — invoice: DejaVuSans 9pt black. insured, term (dashes), due date (+3 weeks).
    p = doc[4]
    _coc_set(p, COC_NAME, company_up, size=9, color=COC_CLR_BLACK, font=COC_FONT_BODY)
    _coc_set(p, "10-16-2025", d["start_dash"], size=9, color=COC_CLR_BLACK, font=COC_FONT_BODY)
    _coc_set(p, "10-16-2026", d["end_dash"], size=9, color=COC_CLR_BLACK, font=COC_FONT_BODY)
    _coc_set(p, "11/6/2025", d["due"], size=9, color=COC_CLR_BLACK, font=COC_FONT_BODY)

    # Page 6 — static boilerplate (no changes). Drop pages 7-12.
    doc.select([0, 1, 2, 3, 4, 5])

    safe = (company_up
            .replace("/","-").replace("\\","-").replace(":","")
            .replace("*","").replace("?","").replace('"',"")
            .replace("<","").replace(">","").replace("|","")
            .replace("'",""))
    out = output_dir / f"COC_{safe}.pdf"
    save_pdf(doc, out)
    doc.close()
    logger.info(f"Confirmation of Coverage saved: {out.name}")
    return out


# ─── COVER WHALE FULL DOCUMENT (all 9 pages) ─────────────────────────────────
# The VIATIC template is a Frankenstein: pages 1-3 are VIATIC LLC (policy
# CUS09114581) but pages 4-9 were spliced in from OTHER quotes — page 4 header
# says CHARLIE HAULING LLC, pages 5-9 say DEKS TRANSPORT LLC, all under policy
# CUS09116580, plus page 6 carries a vehicle/driver schedule and page 7 an
# address — all belonging to those other carriers. generate_coverwhale() makes
# the whole document consistent with the target company.

ALT_POLICY      = "CUS09116580"            # baked-in policy on pages 4-9
ALT_COMPANY_P4  = "CHARLIE HAULING LLC"    # baked-in company on page 4 header
ALT_COMPANY_REST = "DEKS TRANSPORT LLC"    # baked-in company on pages 5-9 headers

# Original schedule values on page 6 / 7 (replaced every run)
CW_OLD_VINS = ["3AKJHHFG3PSNL7399", "1FUJHHDR4NLMZ0512",
               "3AKJHHFG5NSNE9904", "3AKJHHDRXNSNF2060"]
CW_OLD_GARAGE = "17250 DALLAS PKWY, DALLAS, TX 75248"

_MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
           "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
_FIRST_NAMES = ["JAMES", "ROBERT", "MICHAEL", "DAVID", "JOSE", "CARLOS", "JOHN",
                "LUIS", "DANIEL", "ANTHONY", "KEVIN", "BRIAN", "JASON", "ERIC",
                "JUAN", "MARK", "STEVEN", "ANDREW", "RAYMOND", "GREGORY",
                "MIGUEL", "DENNIS", "JERRY", "TYLER", "AARON", "HENRY"]
_LAST_NAMES = ["SMITH", "JOHNSON", "GARCIA", "MARTINEZ", "BROWN", "DAVIS",
               "RODRIGUEZ", "WILSON", "ANDERSON", "THOMAS", "HERNANDEZ", "MOORE",
               "JACKSON", "WHITE", "HARRIS", "CLARK", "LEWIS", "WALKER", "HALL",
               "ALLEN", "YOUNG", "KING", "WRIGHT", "TORRES", "NGUYEN", "REED"]


def _split_full_address(address: str):
    """(street, city, state, zip) from a free-form address string.
    Reuses split_address for the street / city-line split, then peels the zip
    and 2-letter state off the city line."""
    street, rest = split_address(address.upper())
    toks = rest.replace(",", " ").split()
    zipc = state = ""
    if toks and any(c.isdigit() for c in toks[-1]):
        zipc = toks.pop()
    if toks and len(toks[-1]) == 2 and toks[-1].isalpha():
        state = toks.pop()
    city = " ".join(toks)
    return street, city, state, zipc


def _rand_vin(rng):
    chars = "ABCDEFGHJKLMNPRSTUVWXYZ0123456789"   # real VINs omit I, O, Q
    return "".join(rng.choice(chars) for _ in range(17))


def _rand_vehicle_years(rng, n):
    """n years from {2022,2023,2024}, never all identical."""
    pool = [2022, 2023, 2024]
    while True:
        ys = [rng.choice(pool) for _ in range(n)]
        if len(set(ys)) > 1:
            return ys


def _rand_dob(rng):
    """('Mon, DD', 'YYYY') — birth year below 2000."""
    return f"{rng.choice(_MONTHS)}, {rng.randint(1, 28):02d}", str(rng.randint(1965, 1999))


def _rand_hire(rng):
    """('Mon, DD', '2025') — random hire date in 2025."""
    return f"{rng.choice(_MONTHS)}, {rng.randint(1, 28):02d}", "2025"


def _page_spans(page):
    """Flat list of every text span on a page (snapshot of the original layout)."""
    return [s for b in page.get_text("dict")["blocks"]
            for l in b.get("lines", []) for s in l["spans"] if s["text"].strip()]


def _find_span(spans, x0, y0, tol=2.5):
    """Rect of the span whose top-left corner is ~(x0, y0). None if not found."""
    for s in spans:
        bx0, by0, bx1, by1 = s["bbox"]
        if abs(bx0 - x0) <= tol and abs(by0 - y0) <= tol:
            return fitz.Rect(bx0, by0, bx1, by1)
    return None


def _bg_at(pix, x, y, pw, ph):
    """Background color (0-1 floats) of one rendered point. White if no pixmap."""
    if pix is None:
        return (1.0, 1.0, 1.0)
    sx, sy = pix.width / pw, pix.height / ph
    px = max(0, min(pix.width - 1, int(x * sx)))
    py = max(0, min(pix.height - 1, int(y * sy)))
    try:
        return tuple(c / 255.0 for c in pix.pixel(px, py)[:3])
    except Exception:
        return (1.0, 1.0, 1.0)


def _set_rect(page, rect, new_text, *, size, font=FONT_REG, color=(0.0, 0.0, 0.0),
              bg=(1.0, 1.0, 1.0), max_width=None):
    """Cover a span's rect with the row background color and write new_text at
    the same left baseline. Pass bg for shaded (zebra-striped) rows so the
    cover patch doesn't show. Pass max_width to shrink the font when the new
    value would overflow its column (e.g. a wide random VIN crashing into the
    next column)."""
    if rect is None:
        return
    fp = str(font)
    fobj = fitz.Font(fontfile=fp)
    fname = "F" + str(abs(hash(fp)) % 100000)
    if max_width:
        while size > 5.0 and fobj.text_length(new_text, fontsize=size) > max_width:
            size -= 0.2
    # Strip the original span before painting; see _strip_text.
    _strip_text(page, [rect])
    page.draw_rect(fitz.Rect(rect.x0 - 1, rect.y0 - 1, rect.x1 + 1, rect.y1 + 1),
                   color=bg, fill=bg, width=0)
    page.insert_text((rect.x0, rect.y1 - 1.0), new_text,
                     fontfile=fp, fontname=fname, fontsize=size, color=color)


def fill_header_alt(page, old_company, old_policy, company, policy, pix):
    """Top-right header on pages 4-9 (different baked-in company/policy than
    pages 1-3). Right-aligns to the same x=569 edge, DejaVu 7.36 #363636."""
    replace_on_page(page, f"TGL Policy #:  {old_policy}", f"TGL Policy #:  {policy}",
                    fontsize=7.36, top_right_x=569.0, pix=pix, color=CLR_HEADER)
    replace_on_page(page, old_company, company,
                    fontsize=7.36, top_right_x=569.0, pix=pix,
                    x_min=400, y_max=80, color=CLR_HEADER)


def fill_page6_schedule(page, street, city, state, zipc, rng, pix=None):
    """Page 6 — vehicle schedule (VIN + year), garage location, driver schedule
    (names, DOB, hire date). Coordinate-targeted because years/dates repeat.
    The tables are zebra-striped, so each cell is covered with its row's actual
    background color (sampled from pix at a reliably-blank probe point)."""
    spans = _page_spans(page)
    pw, ph = page.rect.width, page.rect.height

    # Vehicle VIN (x0≈48.4) + Year (x0≈156.5) per row. Coordinate-targeted: '2022'
    # repeats 6× on the page. Row bg sampled at x=152 (blank gap before Year).
    # VINs are width-fitted (max_width=103) so a wide random VIN never crashes
    # into the Year column at x=156.5.
    veh_y = [184.7, 206.8, 228.9, 251.0]
    for y, yr in zip(veh_y, _rand_vehicle_years(rng, len(veh_y))):
        row_bg = _bg_at(pix, 152, y, pw, ph)
        _set_rect(page, _find_span(spans, 48.4, y), _rand_vin(rng), size=9.6,
                  bg=row_bg, max_width=103)
        _set_rect(page, _find_span(spans, 156.5, y), str(yr), size=9.6, bg=row_bg)

    # Garage location → company address. Coordinate-targeted (not replace_on_page)
    # because the bold "Garage Location:" label sits flush to the value's left, so
    # left-sampling the bg catches a dark label glyph. Probe a blank spot at x=450.
    _set_rect(page, _find_span(spans, 134.7, 273.1),
              f"{street}, {city}, {state} {zipc}", size=9.6,
              bg=_bg_at(pix, 450, 273.1, pw, ph), max_width=420)

    # Driver schedule rows (size 6.3). Coords from the template's own spans.
    # Row bg probed at x=295 (blank gap between Years-Exp and Date-of-Hire) — the
    # driver table's far-left margin stays white even on grey rows, so we must
    # NOT sample at the left edge.
    rows = [
        dict(fn=(46.4, 494.0), ln=(89.2, 494.0),
             dob_md=(225.3, 490.3), dob_yr=(225.3, 497.8),
             hire_md=(311.0, 490.3), hire_yr=(311.0, 497.8)),
        dict(fn=(46.4, 516.0), ln=(89.2, 516.0),
             dob_md=(225.3, 512.2), dob_yr=(225.3, 519.8),
             hire_md=(311.0, 512.2), hire_yr=(311.0, 519.8)),
        dict(fn=(46.4, 537.9), ln=(89.2, 537.9),
             dob_md=(225.3, 534.2), dob_yr=(225.3, 541.7),
             hire_md=(311.0, 534.2), hire_yr=(311.0, 541.7)),
    ]
    for row in rows:
        row_bg = _bg_at(pix, 295, row["fn"][1], pw, ph)
        fn, ln = rng.choice(_FIRST_NAMES), rng.choice(_LAST_NAMES)
        dob_md, dob_yr = _rand_dob(rng)
        hire_md, hire_yr = _rand_hire(rng)
        vals = {"fn": fn, "ln": ln, "dob_md": dob_md, "dob_yr": dob_yr,
                "hire_md": hire_md, "hire_yr": hire_yr}
        for key, val in vals.items():
            x0, y0 = row[key]
            _set_rect(page, _find_span(spans, x0, y0), val, size=6.3, bg=row_bg)


def fill_page7_address(page, street, city, state, zipc, pix=None):
    """Page 7 — the Address / City / State / Zip row → company address. Each cell
    is width-fitted so a long street/city can't spill into the next column."""
    spans = _page_spans(page)
    pw, ph = page.rect.width, page.rect.height
    row_bg = _bg_at(pix, 250, 115.6, pw, ph)   # blank gap between Address & City
    # (x0, value, max column width before the next column starts)
    cells = [(48.8, street, 273), (327.8, city, 67), (399.5, state, 65), (469.2, zipc, 95)]
    for x0, val, mw in cells:
        _set_rect(page, _find_span(spans, x0, 115.6), val.upper(), size=10.2,
                  bg=row_bg, max_width=mw)


def generate_coverwhale(company: str, usdot: str, address: str, policy: str,
                        output_dir: Path = None, rng=None) -> Path:
    """Fill the WHOLE 9-page Cover Whale policy for one company — pages 1-3 like
    /new, plus pages 4-9 (headers + page-6 schedule + page-7 address) so the
    entire document is consistent with the target company."""
    if output_dir is None:
        output_dir = OUTPUT_DIR
    output_dir.mkdir(exist_ok=True)
    rng = rng or random

    company_up = company.strip().upper()
    addr1, addr2 = split_address(address.upper())
    street, city, state, zipc = _split_full_address(address)

    doc = fitz.open(TEMPLATE_PDF)

    p = doc[0]; pix = p.get_pixmap(dpi=72)
    fill_page1(p, company_up, usdot, addr1, addr2, policy, pix)

    p = doc[1]; pix = p.get_pixmap(dpi=72)
    fill_page2(p, company_up, addr1, addr2, policy, pix)

    p = doc[2]; pix = p.get_pixmap(dpi=72)          # page 3 — VIATIC header
    fill_page_header_only(p, company_up, policy, pix)

    p = doc[3]; pix = p.get_pixmap(dpi=72)          # page 4 — CHARLIE HAULING header
    fill_header_alt(p, ALT_COMPANY_P4, ALT_POLICY, company_up, policy, pix)

    p = doc[4]; pix = p.get_pixmap(dpi=72)          # page 5 — DEKS header
    fill_header_alt(p, ALT_COMPANY_REST, ALT_POLICY, company_up, policy, pix)

    p = doc[5]; pix = p.get_pixmap(dpi=72)          # page 6 — DEKS header + schedule
    fill_header_alt(p, ALT_COMPANY_REST, ALT_POLICY, company_up, policy, pix)
    fill_page6_schedule(p, street, city, state, zipc, rng, pix)

    p = doc[6]; pix = p.get_pixmap(dpi=72)          # page 7 — DEKS header + address
    fill_header_alt(p, ALT_COMPANY_REST, ALT_POLICY, company_up, policy, pix)
    fill_page7_address(p, street, city, state, zipc, pix)

    for i in (7, 8):                                # pages 8 & 9 — DEKS header only
        p = doc[i]; pix = p.get_pixmap(dpi=72)
        fill_header_alt(p, ALT_COMPANY_REST, ALT_POLICY, company_up, policy, pix)

    safe = (company_up
            .replace("/", "-").replace("\\", "-").replace(":", "")
            .replace("*", "").replace("?", "").replace('"', "")
            .replace("<", "").replace(">", "").replace("|", "")
            .replace("'", ""))
    out = output_dir / f"Cover Whale - {safe}.pdf"
    save_pdf(doc, out)
    doc.close()
    logger.info(f"Cover Whale full policy saved: {out.name}")
    return out


# ─── NGANGA DECLARATION PAGE ─────────────────────────────────────────────────

def _ng_fmt_date(d: date) -> str:
    """M/D/YYYY with no leading zeros, matching the template."""
    return f"{d.month}/{d.day}/{d.year}"


def _ng_plus_one_year(d: date) -> date:
    try:
        return d.replace(year=d.year + 1)
    except ValueError:          # Feb 29 -> Feb 28
        return d.replace(year=d.year + 1, day=28)


def _ng_common(**kw):
    """
    Shared options for every Nganga replacement.

    cover=False throughout: the rows on this template are ~13.4pt apart while
    their hit rects are ~14.8pt tall, so a background rectangle inevitably
    bleeds into the neighbouring row — repainting white over the shaded Named
    Insured band, or slicing the value written just above. Redaction already
    removes the old text and leaves the shading untouched, so the rectangle is
    pure downside here.
    """
    return {"fontsize": NG_FONTSIZE, "cover": False,
            "font_reg": CALIBRI_REG, "font_bold": CALIBRI_BOLD,
            "fontname_tag": "cali", **kw}


def _ng_fitted_size(text, fontfile, avail, size=NG_FONTSIZE, min_size=6.5):
    """Largest size up to `size` at which `text` fits within `avail` points."""
    tw = fitz.Font(fontfile=str(fontfile)).text_length(text, fontsize=size)
    return size if tw <= avail else max(min_size, size * avail / tw)


def fill_nganga_header(page, company, policy, pix):
    """Top-right block — identical on both pages, Calibri-Bold, right-aligned."""
    name_line   = f"Named Insured:  {company}"
    policy_line = f"Policy #: {policy}"

    # Both lines grow leftward toward the letterhead logo. Size them together so
    # a long company name never leaves the two header lines mismatched.
    avail = NG_RIGHT_EDGE - NG_LOGO_RIGHT_X
    sz = min(_ng_fitted_size(name_line,   CALIBRI_BOLD, avail),
             _ng_fitted_size(policy_line, CALIBRI_BOLD, avail))

    # These two lines overlap in x, so they must be stripped as one batch — a
    # per-field call would let the policy rect's redaction eat the tail of the
    # company name. y_max=60 keeps the page-1 body row (y196.8) out; it also
    # matches "Named Insured:  <company>".
    replace_many(page, [
        (f"Named Insured:  {NG_COMPANY}", name_line,   dict(y_max=60)),
        (f"Policy #: {NGANGA_POLICY}",    policy_line, dict(y_min=40, y_max=70)),
    ], pix, **_ng_common(fontsize=sz, bold=True, cell_right_x=NG_RIGHT_EDGE))


def fill_nganga_page1(page, company, addr1, addr2, policy, prepared, period, pix):
    """Page 1 — named insured, mailing address, policy period, prepared date."""
    fill_nganga_header(page, company, policy, pix)

    # Value column starts x267.8. Guards bracket the measured hit rects
    # (y196.8–211.5, 210.2–225.0, 223.6–238.4) — they must clear y1, not y0.
    # These rows stack directly on top of one another and share an x-range, so
    # they go through as one batch. merge=True on the dates: their slashes are
    # taller than the digits, so search_for returns each as several fragments.
    replace_many(page, [
        (NG_COMPANY,  company,  dict(y_min=190, y_max=216)),
        (NG_ADDR1,    addr1,    dict(y_min=206, y_max=229)),
        (NG_ADDR2,    addr2,    dict(y_min=219, y_max=242)),
        (NG_PERIOD,   period,   dict(y_min=245, y_max=270, merge=True)),
        (NG_PREPARED, prepared, dict(y_min=706, y_max=735, merge=True)),
    ], pix, **_ng_common(fit_right_x=NG_BOX_RIGHT_X))


def fill_nganga_page2(page, company, policy, vin, year, driver,
                      dob, state, license_no, prepared, pix):
    """Page 2 — covered auto row and covered driver row."""
    fill_nganga_header(page, company, policy, pix)

    # Covered autos (rects y181.5–196.3), column borders x215.8 / 301.2.
    # Covered drivers (rects y277.0–291.8), borders x197.8 / 287.8 / 355.2 / 427.2.
    # The four driver cells share a line, so the whole page goes through as one
    # batch. "FL" also occurs inside the "Florida ..." form names at y402–445,
    # so the licence-state cell is pinned by x as well.
    replace_many(page, [
        (NG_VIN,      vin,        dict(y_min=175, y_max=200, fit_right_x=213.0)),
        (NG_YEAR,     str(year),  dict(y_min=175, y_max=200, fit_right_x=298.0)),
        (NG_DRIVER,   driver,     dict(y_min=270, y_max=296, fit_right_x=195.0)),
        (NG_DOB,      dob,        dict(y_min=270, y_max=296, fit_right_x=352.0,
                                       merge=True)),
        (NG_STATE,    state,      dict(y_min=270, y_max=296, x_min=350,
                                       fit_right_x=424.0)),
        (NG_LICENSE,  license_no, dict(y_min=270, y_max=296, fit_right_x=536.0)),
        (NG_PREPARED, prepared,   dict(y_min=706, y_max=735, merge=True)),
    ], pix, **_ng_common())


def generate_nganga(company: str, address: str, driver: str, policy: str,
                    output_dir: Path = None, today: date = None, rng=None):
    """
    Generate a Motor Carrier Liability Declaration Page.

    company / address  — from All Companies.csv, or entered manually
    driver             — supplied by the user
    policy             — caller-supplied, already incremented

    VIN, model year, DOB, licence state and licence number are randomised.
    Returns (pdf_path, details_dict).
    """
    if output_dir is None:
        output_dir = OUTPUT_DIR
    output_dir.mkdir(exist_ok=True)

    rng     = rng or random
    today   = today or date.today()
    addr1, addr2 = split_address(address)
    company = company.strip()
    driver  = driver.strip()

    # Year first — the VIN encodes it at position 10.
    year       = random_year(rng=rng)
    vin        = random_vin(year, rng=rng)
    dob        = random_dob(today=today, rng=rng)
    state      = random_state(rng=rng)
    last_name  = driver.split()[-1] if driver.split() else ""
    license_no = random_license(state, last_name, rng=rng)

    prepared = _ng_fmt_date(today)
    period   = (f"{_ng_fmt_date(today)} 12:01 AM to "
                f"{_ng_fmt_date(_ng_plus_one_year(today))} 12:00 AM")

    doc = fitz.open(NGANGA_TEMPLATE)

    p = doc[0]; pix = p.get_pixmap(dpi=72)
    fill_nganga_page1(p, company, addr1, addr2, policy, prepared, period, pix)

    p = doc[1]; pix = p.get_pixmap(dpi=72)
    fill_nganga_page2(p, company, policy, vin, year, driver,
                      dob, state, license_no, prepared, pix)

    safe = (company
            .replace("/","-").replace("\\","-").replace(":","")
            .replace("*","").replace("?","").replace('"',"")
            .replace("<","").replace(">","").replace("|","")
            .replace("'",""))
    out = output_dir / f"{safe}_Policy_{today.strftime('%m%d%y')}.pdf"
    save_pdf(doc, out)
    doc.close()

    details = {"policy": policy, "vin": vin, "year": year, "driver": driver,
               "dob": dob, "state": state, "license": license_no,
               "prepared": prepared, "period": period}
    logger.info(f"Nganga saved: {out.name} | {policy} | VIN {vin} ({year}) | "
                f"{driver} {dob} {state} {license_no}")
    return out, details


# ─── SCAN EFFECT ──────────────────────────────────────────────────────────────

def _scan_one(page, dpi: int):
    """Render one PDF page and apply the photo/scan effect. Returns a PIL Image."""
    pix = page.get_pixmap(dpi=dpi)
    img = Image.frombytes("RGB", (pix.width, pix.height), pix.samples)
    w, h = img.size

    # 1. Gray/warm paper tint — blend toward scanner-gray
    paper = Image.new("RGB", img.size, (235, 232, 225))
    img = Image.blend(img, paper, alpha=0.12)

    # 2. Reduce contrast & brightness (washed out / printed look)
    img = ImageEnhance.Contrast(img).enhance(0.82)
    img = ImageEnhance.Brightness(img).enhance(0.93)
    img = ImageEnhance.Sharpness(img).enhance(0.7)

    # 3. Gaussian noise (scanner grain)
    arr = np.array(img, dtype=np.int16)
    noise = np.random.normal(0, 4.5, arr.shape).astype(np.int16)
    arr = np.clip(arr + noise, 0, 255).astype(np.uint8)
    img = Image.fromarray(arr)

    # 4. Blur (scanner/camera softness)
    img = img.filter(ImageFilter.GaussianBlur(radius=0.7))

    # 5. Slight rotation (paper not aligned perfectly)
    angle = random.uniform(-0.7, 0.7)
    img = img.rotate(angle, resample=Image.BICUBIC, expand=False,
                     fillcolor=(230, 228, 222))

    # 6. Subtle edge shadow (very light, not a frame)
    shadow = np.ones((h, w), dtype=np.float32)
    margin_x = int(w * 0.03)
    margin_y = int(h * 0.025)

    for i in range(margin_x):
        f = (i / margin_x) ** 0.8
        shadow[:, i] *= (0.88 + 0.12 * f)
        shadow[:, w - 1 - i] *= (0.90 + 0.10 * f)
    for i in range(margin_y):
        f = (i / margin_y) ** 0.8
        shadow[i, :] *= (0.92 + 0.08 * f)
        shadow[h - 1 - i, :] *= (0.88 + 0.12 * f)

    img_arr = np.array(img, dtype=np.float32)
    for c in range(3):
        img_arr[:, :, c] *= shadow
    img = Image.fromarray(np.clip(img_arr, 0, 255).astype(np.uint8))

    # 7. Slight color temperature shift (warm/yellowish like old scanner)
    final_arr = np.array(img, dtype=np.int16)
    final_arr[:, :, 0] = np.clip(final_arr[:, :, 0] + 3, 0, 255)   # slight red boost
    final_arr[:, :, 2] = np.clip(final_arr[:, :, 2] - 4, 0, 255)   # slight blue drop
    return Image.fromarray(final_arr.astype(np.uint8))


def scannify_pdf(input_path: Path, output_dir: Path = None, dpi: int = 250) -> list[Path]:
    """
    Take a clean PDF and produce JPG images (first 3 pages only)
    that look like photos/scans of a printed document.
    Files are named with the local-time timestamp MMDDYYYYHHMMSS captured
    per page at save time. Returns list of JPG paths.
    """
    if output_dir is None:
        output_dir = input_path.parent

    doc = fitz.open(input_path)
    jpg_paths = []
    num_pages = min(3, len(doc))

    for page_num in range(num_pages):
        img = _scan_one(doc[page_num], dpi)

        # Save as JPG — filename is local-time MMDDYYYYHHMMSS, captured per page.
        # If two pages land in the same second, append _2, _3, ... to avoid overwrite.
        stamp = datetime.now().strftime("%m%d%Y%H%M%S")
        jpg_path = output_dir / f"{stamp}.jpg"
        dup = 2
        while jpg_path.exists():
            jpg_path = output_dir / f"{stamp}_{dup}.jpg"
            dup += 1
        img.save(str(jpg_path), "JPEG", quality=88)
        jpg_paths.append(jpg_path)

    doc.close()
    logger.info(f"Scanned JPGs saved: {[p.name for p in jpg_paths]}")
    return jpg_paths


def scannify_to_pdf(input_path: Path, output_dir: Path = None, dpi: int = 200) -> Path:
    """Scan EVERY page of a PDF and combine them into one scanned-look PDF.
    Each page is JPEG-compressed so the output stays small. Returns the path."""
    if output_dir is None:
        output_dir = input_path.parent

    src = fitz.open(input_path)
    out_doc = fitz.open()
    for i in range(len(src)):
        img = _scan_one(src[i], dpi)
        buf = io.BytesIO()
        img.save(buf, "JPEG", quality=85)
        w, h = img.size
        page = out_doc.new_page(width=w * 72.0 / dpi, height=h * 72.0 / dpi)
        page.insert_image(page.rect, stream=buf.getvalue())
    src.close()

    stamp = datetime.now().strftime("%m%d%Y%H%M%S")
    base = input_path.stem
    if base.startswith("COC_"):           # drop the doc-type prefix from the scan name
        base = base[len("COC_"):]
    out = output_dir / f"{base}_scanned_{stamp}.pdf"
    save_pdf(out_doc, out)
    out_doc.close()
    logger.info(f"Scanned PDF saved: {out.name}")
    return out


# ─── MAIN ────────────────────────────────────────────────────────────────────

def generate():
    print("\n" + "=" * 57)
    print("  Cover Whale PDF Generator  v3")
    print("=" * 57)
    logger.info("=" * 40)
    logger.info("Batch generation started")

    print("\n[1/3] Fonts ...")
    ensure_fonts()
    print("      DejaVuSansCondensed      OK")
    print("      DejaVuSansCondensed-Bold OK")

    print(f"\n[2/3] Reading: {EXCEL_FILE}")
    if not os.path.exists(EXCEL_FILE):
        print("  File not found — check EXCEL_FILE path in script."); sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    hdrs = [str(c.value or "").strip() for c in ws[2]]
    print(f"      Columns : {hdrs}")
    print(f"      Companies: {ws.max_row - 2}")

    def col(*names):
        for n in names:
            for i, h in enumerate(hdrs):
                if h.replace(" ", "").lower() == n.replace(" ", "").lower():
                    return i
        return None

    C_NAME  = col("Legal Name")
    C_USDOT = col("U SDOT Number", "USDOT Number", "USDOT")
    C_ADDR  = col("Physical Address")

    if any(c is None for c in [C_NAME, C_USDOT, C_ADDR]):
        print(f"  Cannot find required columns in: {hdrs}"); sys.exit(1)

    OUTPUT_DIR.mkdir(exist_ok=True)
    print(f"\n[3/3] Generating -> {OUTPUT_DIR}")

    policy  = STARTING_POLICY
    count   = 0
    errors  = []

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not any(row):
            continue

        company = str(row[C_NAME]  or "").strip().upper()
        usdot   = str(row[C_USDOT] or "").strip()
        address = str(row[C_ADDR]  or "").strip().upper()

        if not company:
            continue

        addr1, addr2 = split_address(address)

        print(f"\n  [{count+1:02d}] {company}")
        print(f"        Policy : {policy}")
        print(f"        USDOT  : {usdot}")
        print(f"        Street : {addr1}")
        print(f"        City   : {addr2}")
        logger.info(f"Generating [{count+1:02d}] {company} | Policy: {policy} | USDOT: {usdot}")

        try:
            doc = fitz.open(TEMPLATE_PDF)

            # ── Page 1 ──────────────────────────────────────────────────────
            p   = doc[0]
            pix = p.get_pixmap(dpi=72)   # snapshot BEFORE any changes
            fill_page1(p, company, usdot, addr1, addr2, policy, pix)

            # ── Page 2 ──────────────────────────────────────────────────────
            p   = doc[1]
            pix = p.get_pixmap(dpi=72)
            fill_page2(p, company, addr1, addr2, policy, pix)

            # ── Pages 3+ ────────────────────────────────────────────────────
            for i in range(2, len(doc)):
                p   = doc[i]
                pix = p.get_pixmap(dpi=72)
                fill_page_header_only(p, company, policy, pix)

            # ── Save ────────────────────────────────────────────────────────
            safe = (company
                    .replace("/","-").replace("\\","-").replace(":",  "")
                    .replace("*","").replace("?", "").replace('"', "")
                    .replace("<","").replace(">", "").replace("|",  "")
                    .replace("'",""))
            out = OUTPUT_DIR / f"Cover Whale - {safe}.pdf"
            save_pdf(doc, out)
            doc.close()
            print(f"        Saved  -> {out.name}")
            logger.info(f"Saved: {out.name}")

            policy = increment_policy(policy)
            count += 1

        except Exception as e:
            import traceback
            errors.append((company, str(e)))
            print(f"        ERROR: {e}")
            logger.error(f"Failed: {company} — {e}")
            traceback.print_exc()

    print("\n" + "=" * 57)
    print(f"  Done!  {count} PDFs  ->  {OUTPUT_DIR}")
    if errors:
        print(f"  Errors ({len(errors)}):")
        for n, e in errors:
            print(f"    * {n}: {e}")
    print("=" * 57 + "\n")
    logger.info(f"Batch complete: {count} PDFs, {len(errors)} errors")


if __name__ == "__main__":
    generate()
